from fastapi import FastAPI, Depends, HTTPException, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse
from datetime import timedelta
import os

from models import UserLogin, Token, UserResponse, User
from auth import (
    authenticate_user,
    create_access_token,
    get_user_from_token,
    ACCESS_TOKEN_EXPIRE_MINUTES
)

app = FastAPI(title="API de Autenticación", description="Backend con autenticación JWT")

# Configuración CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:3000", "http://127.0.0.1:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Bearer token para autenticación
security = HTTPBearer()


@app.get("/")
def root():
    """Endpoint de prueba"""
    return {"message": "API de Autenticación funcionando", "status": "OK"}


@app.post("/login", response_model=Token)
def login(user_login: UserLogin):
    """
    Endpoint de login
    - Recibe email y password
    - Valida contra archivo users.txt
    - Retorna token JWT si es válido
    """
    user = authenticate_user(user_login.email, user_login.password)
    
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Credenciales incorrectas",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    # Crear token de acceso
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.email}, expires_delta=access_token_expires
    )
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": user
    }


async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> User:
    """Dependencia para obtener el usuario actual desde el token"""
    token = credentials.credentials
    user = get_user_from_token(token)
    
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token inválido o expirado",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    return user


@app.get("/me", response_model=UserResponse)
def get_current_user_info(current_user: User = Depends(get_current_user)):
    """
    Endpoint protegido que retorna información del usuario autenticado
    Requiere token JWT válido en el header Authorization
    """
    return {
        "id": current_user.id,
        "name": current_user.name,
        "email": current_user.email
    }


# Funciones auxiliares para reportes
def read_users_from_file():
    """Lee los usuarios del archivo users.txt"""
    users = []
    try:
        with open("database/users.txt", "r") as f:
            lines = f.readlines()
            # Saltar la primera línea (encabezado)
            for line in lines[1:]:
                parts = line.strip().split(",")
                if len(parts) >= 4:
                    users.append({
                        "id": parts[0],
                        "name": parts[1],
                        "email": parts[2],
                        "password": parts[3]
                    })
    except FileNotFoundError:
        pass
    return users


def generate_pdf_report():
    """Genera un reporte PDF de usuarios"""
    from fpdf import FPDF
    
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "Reporte de Usuarios", ln=True, align="C")
    pdf.ln(10)
    
    # Encabezado de tabla
    pdf.set_font("Arial", "B", 12)
    pdf.cell(30, 10, "ID", 1)
    pdf.cell(60, 10, "Nombre", 1)
    pdf.cell(80, 10, "Email", 1)
    pdf.ln()
    
    # Datos
    pdf.set_font("Arial", "", 12)
    users = read_users_from_file()
    for user in users:
        pdf.cell(30, 10, user["id"], 1)
        pdf.cell(60, 10, user["name"], 1)
        pdf.cell(80, 10, user["email"], 1)
        pdf.ln()
    
    pdf_path = "database/users_report.pdf"
    pdf.output(pdf_path)
    return pdf_path


def generate_excel_report():
    """Genera un reporte Excel de usuarios"""
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    
    # Encabezados
    ws["A1"] = "ID"
    ws["B1"] = "Nombre"
    ws["C1"] = "Email"
    ws["D1"] = "Contraseña"
    
    # Datos
    users = read_users_from_file()
    for i, user in enumerate(users, start=2):
        ws[f"A{i}"] = user["id"]
        ws[f"B{i}"] = user["name"]
        ws[f"C{i}"] = user["email"]
        ws[f"D{i}"] = user["password"]
    
    excel_path = "database/users_report.xlsx"
    wb.save(excel_path)
    return excel_path


@app.get("/report/pdf")
def get_pdf_report(current_user: User = Depends(get_current_user)):
    """Endpoint para descargar reporte en PDF"""
    try:
        pdf_path = generate_pdf_report()
        return FileResponse(pdf_path, media_type="application/pdf", filename="usuarios_reporte.pdf")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")


@app.get("/report/excel")
def get_excel_report(current_user: User = Depends(get_current_user)):
    """Endpoint para descargar reporte en Excel"""
    try:
        excel_path = generate_excel_report()
        return FileResponse(excel_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="usuarios_reporte.xlsx")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)

